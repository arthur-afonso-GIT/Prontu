"""Leitura e análise segura de planilhas para importação de pacientes."""
import csv
import os
import re
import unicodedata
from datetime import date, datetime


def somente_numeros(valor):
    return "".join(caractere for caractere in str(valor or "") if caractere.isdigit())


def normalizar_texto(valor):
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(caractere for caractere in texto if not unicodedata.combining(caractere))
    return " ".join(texto.casefold().split())


def normalizar_cabecalho(valor):
    return re.sub(r"[^a-z0-9]", "", normalizar_texto(valor))


MAPA_COLUNAS = {
    "nome": "nome", "nomecompleto": "nome", "paciente": "nome",
    "telefone": "telefone", "celular": "telefone", "fone": "telefone", "whatsapp": "telefone",
    "cpf": "cpf", "rg": "rg", "nascimento": "nascimento", "datanascimento": "nascimento", "datadenascimento": "nascimento",
    "convenio": "convenio", "plano": "convenio", "convenioplano": "convenio",
    "pasta": "pasta", "grupo": "pasta", "especialidade": "pasta",
    "sexo": "sexo", "sexobiologico": "sexo", "estadocivil": "estado_civil",
    "profissao": "profissao", "endereco": "endereco", "enderecoresidencial": "endereco",
    "queixa": "queixa", "queixaprincipal": "queixa",
}

CAMPOS = ("nome", "telefone", "cpf", "rg", "nascimento", "convenio", "pasta", "sexo",
          "estado_civil", "profissao", "endereco", "queixa")


def _texto(valor):
    return str(valor or "").strip()


def normalizar_data(valor):
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")
    texto = _texto(valor)
    if not texto:
        return ""
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(texto, formato).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def ler_planilha(caminho):
    """Retorna cabeçalhos e linhas cruas de CSV ou XLSX."""
    extensao = os.path.splitext(caminho)[1].casefold()
    if extensao == ".csv":
        ultimo_erro = None
        for codificacao in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                with open(caminho, "r", encoding=codificacao, newline="") as arquivo:
                    amostra = arquivo.read(4096)
                    arquivo.seek(0)
                    try:
                        dialecto = csv.Sniffer().sniff(amostra, delimiters=";,\t")
                    except csv.Error:
                        dialecto = csv.excel
                    leitor = csv.reader(arquivo, dialect=dialecto)
                    linhas = [linha for linha in leitor if any(_texto(celula) for celula in linha)]
                    break
            except UnicodeDecodeError as erro:
                ultimo_erro = erro
        else:
            raise ValueError("Não foi possível ler a codificação do CSV.") from ultimo_erro
    elif extensao == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as erro:
            raise ValueError("A leitura de Excel requer a biblioteca openpyxl.") from erro
        livro = load_workbook(caminho, read_only=True, data_only=True)
        planilha = livro.active
        linhas = [list(linha) for linha in planilha.iter_rows(values_only=True)
                  if any(_texto(celula) for celula in linha)]
        livro.close()
    else:
        raise ValueError("Escolha um arquivo CSV ou Excel (.xlsx).")

    if len(linhas) < 2:
        raise ValueError("A planilha precisa ter cabeçalho e ao menos um paciente.")
    return linhas[0], linhas[1:]


def preparar_registros(cabecalhos, linhas):
    """Relaciona colunas conhecidas aos campos do Prontu e valida nomes."""
    colunas = {}
    for indice, titulo in enumerate(cabecalhos):
        campo = MAPA_COLUNAS.get(normalizar_cabecalho(titulo))
        if campo and campo not in colunas:
            colunas[campo] = indice
    if "nome" not in colunas:
        raise ValueError("Não encontramos uma coluna de nome. Use 'Nome', 'Nome completo' ou 'Paciente'.")

    registros, erros = [], []
    for numero_linha, linha in enumerate(linhas, start=2):
        dados = {campo: "" for campo in CAMPOS}
        for campo, indice in colunas.items():
            dados[campo] = _texto(linha[indice]) if indice < len(linha) else ""
        dados["cpf"] = somente_numeros(dados["cpf"])[:11]
        dados["telefone"] = somente_numeros(dados["telefone"])
        dados["nascimento"] = normalizar_data(dados["nascimento"])
        if not dados["nome"]:
            erros.append(f"Linha {numero_linha}: nome não informado.")
            continue
        dados["_linha"] = numero_linha
        registros.append(dados)
    return registros, set(colunas), erros


def classificar_registros(registros, existentes):
    """Classifica cada registro como novo ou existente dentro da clínica."""
    por_cpf, por_nome_telefone = {}, {}
    for existente in existentes:
        cpf = somente_numeros(existente.get("cpf"))[:11]
        if cpf:
            por_cpf[cpf] = existente
        nome = normalizar_texto(existente.get("nome"))
        telefone = somente_numeros(existente.get("telefone"))
        if nome and telefone:
            por_nome_telefone[(nome, telefone)] = existente

    resultado = []
    vistos = set()
    for registro in registros:
        cpf = registro["cpf"]
        chave_nome_telefone = (normalizar_texto(registro["nome"]), registro["telefone"])
        existente = por_cpf.get(cpf) if cpf else None
        if not existente and all(chave_nome_telefone):
            existente = por_nome_telefone.get(chave_nome_telefone)
        chave_linha = (cpf or "",) + chave_nome_telefone
        duplicado_no_arquivo = chave_linha in vistos and any(chave_linha)
        vistos.add(chave_linha)
        resultado.append({"dados": registro, "existente": existente, "duplicado_no_arquivo": duplicado_no_arquivo})
    return resultado


def payload_novo(registro, consultorio_id):
    return {
        "consultorio_id": consultorio_id,
        "nome": registro["nome"],
        "telefone": registro["telefone"],
        "nascimento": registro["nascimento"] or "1990-01-01",
        "convenio": registro["convenio"] or "PARTICULAR",
        "pasta": registro["pasta"] or "Geral",
        "sexo": registro["sexo"] or "Masculino",
        "cpf": registro["cpf"],
        "rg": registro["rg"],
        "estado_civil": registro["estado_civil"],
        "profissao": registro["profissao"],
        "endereco": registro["endereco"],
        "queixa": registro["queixa"],
    }


def payload_atualizacao(registro, campos_presentes):
    campos_banco = {"nome", "telefone", "nascimento", "convenio", "pasta", "sexo", "cpf", "rg",
                     "estado_civil", "profissao", "endereco", "queixa"}
    return {campo: registro[campo] for campo in campos_presentes & campos_banco if registro.get(campo) != ""}
